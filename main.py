import openpyxl
import itertools

fileName = input('Name of file you want to input: ')
wb = openpyxl.load_workbook(fileName)

wb = openpyxl.load_workbook('TestingData (1NF-5NF).xlsx')
wb = openpyxl.load_workbook('TestingData (5NF violation).xlsx')

sheet = wb.active

tables = []

# Parses from xlsx file by checking for borders to indicate a cell of a table and following the border to trace the table.

tableCount = -1
prevCoord = -1
tables.append([])
rowCount = 0
for cell in sheet['A']:
    if cell.border.top.style or cell.border.left.style or cell.border.right.style or cell.border.bottom.style:
        row = cell.coordinate[1:]
        if cell.coordinate[1:] != '1':
            if int(prevCoord) != int(row) - 1:
                tableCount += 1
                tables.append([])
                tables[tableCount].append([])
                rowCount = 0
            else:
                tables[tableCount].append([])
                rowCount += 1
        for nest1 in sheet[row]:
            if nest1.value != None:
                tables[tableCount][rowCount].append(nest1.value)
        prevCoord = cell.coordinate[1:]

normalFormChoice = str(input('Normal form level to normalize to (integer or \'bcnf\' for bcnf): '))

# 1nf ---------------------------------------------------

if normalFormChoice == '1' or normalFormChoice == '2' or normalFormChoice == '3' or normalFormChoice == '4' or normalFormChoice == '5' or normalFormChoice == "bcnf":
    # Inputs primary keys
    tableChoice = int(input('which table to normalize?'))
    primaryKey = input('Primary Key (type \'done\' when finished): ')
    primaryKeys = [primaryKey]
    while primaryKey != 'done':
        primaryKey = input('Primary Key (type \'done\' when finished): ')
        if primaryKey != 'done':
            primaryKeys.append(primaryKey)

    originalPKs = primaryKeys.copy()

    primaryKeyx = []
    for value in primaryKeys:
        primaryKeyx.append(tables[tableChoice][0].index(value))

    # Identifies multivalued coords by looking for curly braces.
    multivaluedxCoords = []
    for i in range(len(tables[tableChoice])):
        for j in range(len(tables[tableChoice][i])):
            if '{' in str(tables[tableChoice][i][j]) and '}' in str(tables[tableChoice][i][j]):
                if j not in multivaluedxCoords:
                    multivaluedxCoords.append(j)
    multivaluedxCoords.sort(key=lambda x: x)

    # Creates new tables for 1nf and populates them.
    firstNFTables = []
    firstNFTableCount = -1
    for i in multivaluedxCoords:
        row = 1
        pullRow = 1
        firstNFTableCount += 1
        firstNFTables.append([])
        firstNFTables[firstNFTableCount].append([])
        firstNFTables[firstNFTableCount][0].extend(primaryKeys)
        firstNFTables[firstNFTableCount][0].append(tables[tableChoice][0][i])
        for j in range(1, len(tables[tableChoice])):
            if '{' in str(tables[tableChoice][j][i]) and '}' in str(tables[tableChoice][j][i]):
                multivalues = tables[tableChoice][j][i][tables[tableChoice][j][i].index('{')+1:tables[tableChoice][j][i].index('}')].split(', ')
                for k in multivalues:
                    firstNFTables[firstNFTableCount].append([])
                    for l in primaryKeyx:
                        firstNFTables[firstNFTableCount][row].append(tables[tableChoice][pullRow][l])
                    firstNFTables[firstNFTableCount][row].append(k)
                    row += 1
            else:
                firstNFTables[firstNFTableCount].append([])
                for l in primaryKeyx:
                    firstNFTables[firstNFTableCount][row].append(tables[tableChoice][pullRow][l])
                firstNFTables[firstNFTableCount][row].append(tables[tableChoice][pullRow][i])
                row += 1
            pullRow += 1


    multivaluedAtts = []
    for i in multivaluedxCoords:
        if i not in multivaluedAtts:
            multivaluedAtts.append(i)

    # Populates original table once all multivalued attributes are removed.
    firstNFTables.append([])
    firstNFTableCount += 1
    for i in range(len(tables[tableChoice])):
        firstNFTables[firstNFTableCount].append([])
        for j in range(len(tables[tableChoice][i])):
            if j not in multivaluedAtts:
                firstNFTables[firstNFTableCount][i].append(tables[tableChoice][i][j])

    primaryKeysAfter1nf = []

    for i in range(len(firstNFTables) - 1):
        primaryKeysAfter1nf.append(firstNFTables[i][0])

    primaryKeysAfter1nf.append(primaryKeys)

    print('----------')
    print('1nf:')
    for i in firstNFTables:
        print('----------')
        for j in i:
            print(j)
        print('Primary keys:', primaryKeysAfter1nf[firstNFTables.index(i)])

# 2nf ---------------------------------------------------

if normalFormChoice == '2' or normalFormChoice == '3' or normalFormChoice == '4' or normalFormChoice == '5' or normalFormChoice == "bcnf":
    secondNFTables = firstNFTables.copy()
    primaryKeysAfter2nf = primaryKeysAfter1nf.copy()
    appendCount = len(secondNFTables)

    # Identifies and parses functinoal dependencies from input.
    FDLeftList = []
    FDRightList = []
    FD = ''
    while FD != 'done':
        FD = input("Functional Dependency (\'done\' if done): ")
        if FD == 'done':
            break
        splitFD = FD.split(' -> ')
        if '{' in splitFD[0] and '}' in splitFD[0]:
            FDLeftList.append(splitFD[0][splitFD[0].index('{')+1:splitFD[0].index('}')].split(', '))
        else:
            FDLeftList.append(splitFD[0].split(', '))
        if '{' in splitFD[1] and '}' in splitFD[1]:
            FDRightList.append(splitFD[1][splitFD[1].index('{')+1:splitFD[1].index('}')].split(', '))
        else:
            FDRightList.append(splitFD[1].split(', '))

    for i in range(len(firstNFTables)):
        for j in range(len(FDLeftList)):
            # First checks if all attributes are in table, then checks if left matches primary keys, then checks if any attribute exists in primary key.
            isIn = True
            for k in FDLeftList[j]:
                if k not in firstNFTables[i][0]:
                    isIn = False
                    break
            if isIn == False:
                continue
            for k in FDRightList[j]:
                if k not in firstNFTables[i][0]:
                    isIn = False
                    break
            if isIn == False:
                continue
            if primaryKeysAfter1nf[i] == FDLeftList[j]:
                continue
            isIn = False
            for k in FDLeftList[j]:
                if k in primaryKeysAfter1nf[i]:
                    isIn = True
                    break
            if isIn == False:
                continue
            secondNFTables.append([])
            primaryKeysAfter2nf.append([])
            for k in range(len(firstNFTables[i])):
                secondNFTables[appendCount].append([])
            for k in FDLeftList[j]:
                indexToAdd = firstNFTables[i][0].index(k)
                primaryKeysAfter2nf[appendCount].append(k)
                for l in range(len(secondNFTables[appendCount])):
                    secondNFTables[appendCount][l].append(secondNFTables[i][l][indexToAdd])
            for k in FDRightList[j]:
                indexToAdd = firstNFTables[i][0].index(k)
                for l in range(len(secondNFTables[appendCount])):
                    pop2nf = secondNFTables[i][l].pop(indexToAdd)
                    secondNFTables[appendCount][l].append(pop2nf)
            # Checks if there are redundant tuples after moving to new table. Removes redundant tuples.
            removeRedundancyFromAppend = []
            [removeRedundancyFromAppend.append(x) for x in secondNFTables[appendCount] if x not in removeRedundancyFromAppend]
            secondNFTables.pop(appendCount)
            secondNFTables.append(removeRedundancyFromAppend)
            if secondNFTables[i][0] == primaryKeysAfter2nf[i]:
                secondNFTables.pop(i)
                primaryKeysAfter2nf.pop(i)
            appendCount += 1

    print('----------')
    print('2nf:')
    for i in secondNFTables:
        print('----------')
        for j in i:
            print(j)
        print('Primary keys:', primaryKeysAfter2nf[secondNFTables.index(i)])

# 3nf ---------------------------------------------------

# Check if all attributes are in table, then check if left has no primary keys, then loop through other FDs, check if all their values are present in table, if they are, check if left of first is present in right of second. If yes, split.
# j (outer) makes sure left values aren't in primary key. k (inner) makes sure left values are in pk

if normalFormChoice == '3':
    thirdNFTables = secondNFTables.copy()
    primaryKeysAfter3nf = primaryKeysAfter2nf.copy()
    appendCount = len(thirdNFTables)

    # Performs normalization on any functional dependencies that don't have left side as superkey or right side as prime attribute.
    for i in thirdNFTables:
        for j in range(len(FDLeftList)):
            isIn = True
            for k in FDLeftList[j]:
                if k not in thirdNFTables[thirdNFTables.index(i)][0]:
                    isIn = False
                    break
            if isIn == False:
                continue
            for k in FDRightList[j]:
                if k not in thirdNFTables[thirdNFTables.index(i)][0]:
                    isIn = False
                    break
            if isIn == False:
                continue
            if primaryKeysAfter3nf[thirdNFTables.index(i)] == FDLeftList[j]:
                continue
            isIn = False
            for k in FDRightList[j]:
                if k in primaryKeysAfter3nf[thirdNFTables.index(i)]:
                    isIn = True
                    break
            if isIn == True:
                continue
            for k in range(len(FDLeftList)):
                if k == j:
                    continue
                isIn = True
                for l in FDLeftList[k]:
                    if l not in thirdNFTables[thirdNFTables.index(i)][0]:
                        isIn = False
                        break
                if isIn == False:
                    continue
                for l in FDRightList[k]:
                    if l not in thirdNFTables[thirdNFTables.index(i)][0]:
                        isIn = False
                        break
                if isIn == False:
                    continue
                if primaryKeysAfter3nf[thirdNFTables.index(i)] != FDLeftList[k]:
                    continue
                isIn = True
                for l in FDLeftList[j]:
                    if l not in FDRightList[k]:
                        isIn = False
                        break
                if isIn == False:
                    continue
                thirdNFTables.append([])
                primaryKeysAfter3nf.append([])
                for l in range(len(thirdNFTables[thirdNFTables.index(i)])):
                    thirdNFTables[appendCount].append([])
                for l in FDLeftList[j]:
                    indexToAdd = thirdNFTables[thirdNFTables.index(i)][0].index(l)
                    primaryKeysAfter3nf[appendCount].append(l)
                    for m in range(len(thirdNFTables[appendCount])):
                        thirdNFTables[appendCount][m].append(thirdNFTables[thirdNFTables.index(i)][m][indexToAdd])
                for l in FDRightList[j]:
                    indexToAdd = thirdNFTables[thirdNFTables.index(i)][0].index(l)
                    for m in range(len(thirdNFTables[appendCount])):
                        pop3nf = secondNFTables[thirdNFTables.index(i)][m].pop(indexToAdd)
                        thirdNFTables[appendCount][m].append(pop3nf)
                appendCount += 1

    print('----------')
    print('3nf:')
    for i in thirdNFTables:
        print('----------')
        for j in i:
            print(j)
        print('Primary keys:', primaryKeysAfter3nf[thirdNFTables.index(i)])
    print(len(primaryKeysAfter3nf))

# bcnf ---------------------------------------------------

if normalFormChoice == '4' or normalFormChoice == '5' or normalFormChoice == "bcnf":
    thirdNFTables = secondNFTables.copy()
    primaryKeysAfter3nf = primaryKeysAfter2nf.copy()
    appendCount = len(thirdNFTables)

    # Performs normalization on any functional dependencies that don't have left side as superkey.
    for i in thirdNFTables:
        for j in range(len(FDLeftList)):
            isIn = True
            for k in FDLeftList[j]:
                if k not in thirdNFTables[thirdNFTables.index(i)][0]:
                    isIn = False
                    break
            if isIn == False:
                continue
            for k in FDRightList[j]:
                if k not in thirdNFTables[thirdNFTables.index(i)][0]:
                    isIn = False
                    break
            if isIn == False:
                continue
            if primaryKeysAfter3nf[thirdNFTables.index(i)] == FDLeftList[j]:
                continue
            for k in range(len(FDLeftList)):
                if k == j:
                    continue
                isIn = True
                for l in FDLeftList[k]:
                    if l not in thirdNFTables[thirdNFTables.index(i)][0]:
                        isIn = False
                        break
                if isIn == False:
                    continue
                for l in FDRightList[k]:
                    if l not in thirdNFTables[thirdNFTables.index(i)][0]:
                        isIn = False
                        break
                if isIn == False:
                    continue
                if primaryKeysAfter3nf[thirdNFTables.index(i)] != FDLeftList[k]:
                    continue
                isIn = True
                for l in FDLeftList[j]:
                    if l not in FDRightList[k]:
                        isIn = False
                        break
                if isIn == False:
                    continue
                thirdNFTables.append([])
                primaryKeysAfter3nf.append([])
                for l in range(len(thirdNFTables[thirdNFTables.index(i)])):
                    thirdNFTables[appendCount].append([])
                for l in FDLeftList[j]:
                    indexToAdd = thirdNFTables[thirdNFTables.index(i)][0].index(l)
                    primaryKeysAfter3nf[appendCount].append(l)
                    for m in range(len(thirdNFTables[appendCount])):
                        thirdNFTables[appendCount][m].append(thirdNFTables[thirdNFTables.index(i)][m][indexToAdd])
                for l in FDRightList[j]:
                    indexToAdd = thirdNFTables[thirdNFTables.index(i)][0].index(l)
                    for m in range(len(thirdNFTables[appendCount])):
                        pop3nf = secondNFTables[thirdNFTables.index(i)][m].pop(indexToAdd)
                        thirdNFTables[appendCount][m].append(pop3nf)
                appendCount += 1

    print('----------')
    print('bcnf:')
    for i in thirdNFTables:
        print('----------')
        for j in i:
            print(j)
        print('Primary keys:', primaryKeysAfter3nf[thirdNFTables.index(i)])

# 4nf ---------------------------------------------------

if normalFormChoice == '4' or normalFormChoice == '5':
    fourthNFTables = thirdNFTables.copy()
    primaryKeysAfter4nf = primaryKeysAfter3nf.copy()
    appendCount = len(thirdNFTables)

    # Inputs and parses multivalued dependencies.
    mvdLeftList = []
    mvdRightList = []
    mvd = ''
    while mvd != 'done':
        mvd = input("Multivalued Dependency (\'done\' if done): ")
        if mvd == 'done':
            break
        splitmvd = mvd.split(' ->> ')
        if '{' in splitmvd[0] and '}' in splitmvd[0]:
            mvdLeftList.append(splitmvd[0][splitmvd[0].index('{')+1:splitmvd[0].index('}')].split(', '))
        else:
            mvdLeftList.append(splitmvd[0].split(', '))
        mvdRightList.append(splitmvd[1].split('|'))

    # Checks tables for presence of all attributes of a dependency and splits table while removing redundancies.
    for i in range(len(fourthNFTables)):
        for j in range(len(mvdLeftList)):
            isIn = True
            for k in mvdLeftList[j]:
                if k not in fourthNFTables[i][0]:
                    isIn = False
                    break
            if isIn == False:
                continue
            for k in mvdRightList[j]:
                if k not in fourthNFTables[i][0]:
                    isIn = False
                    break
            if isIn == False:
                continue
            for k in mvdRightList[j]:
                fourthNFTables.append([])
                primaryKeysAfter4nf.append([])
                for l in range(len(fourthNFTables[i])):
                    fourthNFTables[appendCount].append([])
                for l in mvdLeftList[j]:
                    indexToAdd = fourthNFTables[i][0].index(l)
                    primaryKeysAfter4nf[appendCount].append(l)
                    for m in range(len(fourthNFTables[appendCount])):
                        fourthNFTables[appendCount][m].append(fourthNFTables[i][m][indexToAdd])
                indexToAdd = fourthNFTables[i][0].index(k)
                primaryKeysAfter4nf[appendCount].append(k)
                for l in range(len(fourthNFTables[appendCount])):
                    pop4nf = fourthNFTables[i][l].pop(indexToAdd)
                    fourthNFTables[appendCount][l].append(pop4nf)
                removeRedundancyFromAppend = []
                [removeRedundancyFromAppend.append(x) for x in fourthNFTables[appendCount] if x not in removeRedundancyFromAppend]
                fourthNFTables.pop(appendCount)
                fourthNFTables.append(removeRedundancyFromAppend)
                if fourthNFTables[appendCount][0] == originalPKs:
                    fourthNFTables.pop(appendCount)
                    primaryKeysAfter4nf.pop(appendCount)
                else:
                    appendCount += 1
            if fourthNFTables[i][0] == primaryKeysAfter4nf[i]:
                fourthNFTables.pop(i)
                primaryKeysAfter4nf.pop(i)
                appendCount -= 1

    print('----------')
    print('4nf:')
    for i in fourthNFTables:
        print('----------')
        for j in i:
            print(j)
        print('Primary keys:', primaryKeysAfter4nf[fourthNFTables.index(i)])

# 5nf ---------------------------------------------------

if normalFormChoice == '5':
    fifthNFTables = fourthNFTables.copy()
    primaryKeysAfter5nf = primaryKeysAfter4nf.copy()
    tablesToPop = []

    for i in range(len(fourthNFTables)):
        matchFound = False
        possibleCommonAttributes = []
        for r in range(1,len(fourthNFTables[i][0]) - 1):
            possibleCommonAttributes.extend(itertools.combinations(fourthNFTables[i][0], r))
        for j in possibleCommonAttributes:
            remainingAttributes = []
            for k in fourthNFTables[i][0]:
                if k not in j:
                    remainingAttributes.append(k)
            possibleRemainingLeft = []
            for r in range(1,len(remainingAttributes)):
                possibleRemainingLeft.extend(itertools.combinations(remainingAttributes, r))
            for k in possibleRemainingLeft:
                possibleRemainingRight = []
                for l in fourthNFTables[i][0]:
                    if l not in k and l not in j:
                        possibleRemainingRight.append(l)
                # populates left table and right table, then check if union yields original table.
                leftTable = []
                rightTable = []
                for l in range(len(fourthNFTables[i])):
                    leftTable.append([])
                    rightTable.append([])
                for l in j:
                    indexToAdd = fourthNFTables[i][0].index(l)
                    for m in range(len(leftTable)):
                        leftTable[m].append(fourthNFTables[i][m][indexToAdd])
                        rightTable[m].append(fourthNFTables[i][m][indexToAdd])
                for l in k:
                    indexToAdd = fourthNFTables[i][0].index(l)
                    for m in range(len(leftTable)):
                        leftTable[m].append(fourthNFTables[i][m][indexToAdd])
                for l in possibleRemainingRight:
                    indexToAdd = fourthNFTables[i][0].index(l)
                    for m in range(len(rightTable)):
                        rightTable[m].append(fourthNFTables[i][m][indexToAdd])
                removeRedundancyFromAppend = []
                [removeRedundancyFromAppend.append(x) for x in leftTable if x not in removeRedundancyFromAppend]
                leftTable = removeRedundancyFromAppend.copy()
                removeRedundancyFromAppend = []
                [removeRedundancyFromAppend.append(x) for x in rightTable if x not in removeRedundancyFromAppend]
                rightTable = removeRedundancyFromAppend.copy()
                newTable = []
                appendCount2 = 0
                commonAttributesStart = len(j)
                for l in leftTable:
                    for m in rightTable:
                        if l[:commonAttributesStart] == m[:commonAttributesStart]:
                            newTable.append([])
                            newTable[appendCount2].extend(l)
                            newTable[appendCount2].extend(m[commonAttributesStart:])
                            appendCount2 += 1
                if newTable == fourthNFTables[i]:
                    matchFound = True
                    break
            if matchFound == True:
                break
        fifthNFTables.append(leftTable)
        fifthNFTables.append(rightTable)
        leftpks = []
        rightpks = []
        for j in primaryKeysAfter5nf[i]:
            if j in leftTable[0]:
                leftpks.append(j)
            if j in rightTable[0]:
                rightpks.append(j)
        primaryKeysAfter5nf.append(leftpks)
        primaryKeysAfter5nf.append(rightpks)
        if matchFound == True:
            tablesToPop.append(i)
    
    for i in tablesToPop:
        fifthNFTables.pop(i)
        primaryKeysAfter5nf.pop(i)

    print('----------')
    print('5nf:')
    for i in fifthNFTables:
        print('----------')
        for j in i:
            print(j)
        print('Primary keys:', primaryKeysAfter5nf[fifthNFTables.index(i)])
